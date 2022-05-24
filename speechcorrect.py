# coding=utf-8


import pinyin
import pandas
import openpyxl


def read_text(file_path):                               #获取excel文档信息函数
    excel_df = pandas.read_excel(file_path)                #读取excel文件
    text_series = list(excel_df['Text'])
    vedioID_series = list(excel_df['VideoID'])            #读取文件中各列信息
    StartTime = list(excel_df['StartTime'])
    EndTime = list(excel_df['EndTime'])
    return vedioID_series,StartTime,EndTime,text_series              #返回视频编号代码，和各行识别内容


def read_dic(file_path_list):       #读取正确数据集作为匹配标准
    rightwords = []
    rightpinyins = []
    for fp in file_path_list:
        with open(fp, 'r', encoding='utf-8') as f:
            for line in f:
                if line == '\n':
                    line = "夏黎明"
                info = line.split()
                rightwords.append(info[0])      #读取汉字内容
                rightpinyins.append(pinyin.get(info[0],format='strip'))     #中文文本转换为拼音
    rightwords = list(set(rightwords))
    return rightwords                       #返回标准中文文本、标准拼音文本


def judge_error(first_strs, WordInDoubt, PotentialFix, files):    #判断错误函数
    count = 0
    rightwords = read_dic(file_path_list=files)     #调用函数读取标准词
    rightwords = list(set(rightwords))              #去除列表中重复标准词
    rightpinyins = []
    for rightword in rightwords:
        rightpinyins.append(pinyin.get(rightword, format="strip"))  #标准词转换为拼音
    strs = list(first_strs)
    wordstr = ""
    strpinyin = ""
    for k in range(len(rightwords)):            #避免被修改句子长度短于标准词长度，增加其长度
        if len(strs) < len(rightwords[k]):
            for n in range(99):
                strs.append('日')                #在较短被修改句子后面增加字，以增长句子长度
                if len(strs) == len(rightwords[k]):
                    break
        for i in range(len(strs)):
            if i == len(strs) - (len(rightwords[k]) - 1):       #避免读取的范围超过列表长度
                break
            # 从转换语音文件的一行中逐一截取与匹配标准词等长度的词
            for j in range(len(rightwords[k])):
                wordstr = wordstr + strs[i + j]
            strpinyin = pinyin.get(wordstr, format="strip") #将逐一截取的词转化为拼音
            if rightpinyins[k] == strpinyin and wordstr != rightwords[k]:    #如果截取的词拼音与标准词一样则列为怀疑词
                count = count + 1
                WordInDoubt.append(wordstr)
                PotentialFix.append(rightwords[k])
            wordstr = ""
            strpinyin = ""
    return count,WordInDoubt,PotentialFix


def write_excel(VD,Tt,Wt,Px,ST,ET,filename):       #写入excel文件函数
    workbook = openpyxl.Workbook()
    sheet = workbook.create_sheet("Corrector")      #创建表单corrector
    arrlist = ["VideoID","StartTime","EndTime","Text","WordInDoubt","PotentialFix"]
    for i in range(len(arrlist)):       #写入每列标题
        sheet.cell(1,i+1,arrlist[i])
    for i in range(len(VD)):        #写入每列对应内容（视频序号、对应句子、怀疑词汇、纠正词汇）
        sheet.cell(i+2,1,VD[i])
        sheet.cell(i+2,2,ST[i])
        sheet.cell(i+2,3,ET[i])
        sheet.cell(i+2,4,Tt[i])
        sheet.cell(i+2,5,Wt[i])
        sheet.cell(i+2,6,Px[i])
    workbook.save(filename)         #保存数据表名为"save_error"


def corrector(files,file_path,filename):
    WordInDoubt = []
    PotentialFix = []
    VideoID, Text =[],[]
    ST, ET = [],[]
    # for i in range(57-47):              #构造标准词与待修改文件的文件名
    #     file_path = "D:\\GraduationCorrectDic\\generate\\speechcontent_"+str(47+i)+".xlsx"
    #     files = ["D:\\GraduationCorrectDic\\generate\\slideskeywords_"+str(47+i)+".txt",
    #              "D:\\GraduationCorrectDic\\generate\\labels_"+str(47+i)+".txt"]
    ID,StartTime,EndTime,willcorrectwords = read_text(file_path=file_path)    #调用函数读取每行视频序号与每行Text内容
    for i in range(len(willcorrectwords)):
        #判断每行错误词
        count, Wt, Px= judge_error(willcorrectwords[i],WordInDoubt,PotentialFix,files)
        for j in range(count):                  #获取存在错误句子的视频序号与句子内容
            VideoID.append(str(ID[i]))
            ST.append(str(StartTime[i]))
            ET.append(str(EndTime[i]))
            Text.append(str(willcorrectwords[i]))
    write_excel(VideoID,Text,Wt,Px,ST,ET,filename)             #写入excel中
    print("修改文件已保存")

















